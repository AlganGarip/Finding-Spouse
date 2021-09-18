from openpyxl import Workbook,load_workbook 
wb=load_workbook("religion-survey-results.xlsx")
ws = wb.active
def main():
    print('Welcome to the program for finding spouse')
    wanted_gender=input('Which gender do you want to marry? ')
    wanted_country=input('Which contry do you want to marry from? ')
    wanted_religion=input('Which religious state do you want your'+
                          ' spouse to have?(Muslim etc.) ')
    number_of_suitables=0
    for row in range(3,1042):
        gender=str(ws.cell(row,46).value)
        religion=str(ws.cell(row,1).value)  
        country=str(ws.cell(row,48).value)
        if gender==wanted_gender and country==wanted_country and religion==wanted_religion:
            number_of_suitables+=1
    if number_of_suitables==0:
        print('You are unlucky.There is ',number_of_suitables,'candidate')
    elif number_of_suitables<=5:
        print('You have to be in hurry.There is ',number_of_suitables,'candidates')
    else:
        print('You are lucky.There is ',number_of_suitables,'candidates')
main()